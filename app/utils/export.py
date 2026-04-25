"""Export utilities — Excel and PDF generation."""
import io
from datetime import datetime


def export_to_excel(data, headers, sheet_name='تقرير', filename_prefix='report'):
    """Generate an Excel file from a list of dicts."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from flask import make_response

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.sheet_view.rightToLeft = True

        # Header styling
        header_fill = PatternFill(start_color='006C49', end_color='006C49', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_align = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[chr(64 + col_idx)].width = 20

        # Write data
        for row_idx, row in enumerate(data, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='right')
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F4F6', end_color='F2F4F6', fill_type='solid')

        ws.row_dimensions[1].height = 30

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename_prefix}_{timestamp}.xlsx'
        return response

    except ImportError:
        from flask import flash, redirect, url_for
        flash('مكتبة openpyxl غير متوفرة. قم بتثبيتها أولاً.', 'error')
        return redirect(url_for('dashboard.index'))


def export_to_pdf(html_content, filename_prefix='report'):
    """Generate a PDF from HTML content."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from flask import make_response

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        styles = getSampleStyleSheet()

        elements = []
        # Basic paragraph from HTML (simplified)
        elements.append(Paragraph(html_content, styles['Normal']))
        doc.build(elements)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={filename_prefix}_{timestamp}.pdf'
        return response

    except ImportError:
        from flask import flash, redirect, url_for
        flash('مكتبة reportlab غير متوفرة.', 'error')
        return redirect(url_for('dashboard.index'))
